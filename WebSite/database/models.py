"""
Database Models and Schema for Insurance Comparator
"""

import mysql.connector
from mysql.connector import pooling
from datetime import datetime
from typing import Dict, Any, List, Optional
import json
import os
from urllib.parse import urlparse

# Parse MySQL configuration from environment
def get_db_config():
    """Get database configuration from environment variables"""
    # Check if MYSQL_URL is provided (Railway format)
    mysql_url = os.getenv('MYSQL_URL')

    if mysql_url:
        # Parse the MySQL URL
        parsed = urlparse(mysql_url)
        return {
            'host': parsed.hostname or 'localhost',
            'port': parsed.port or 3306,
            'user': parsed.username or 'root',
            'password': parsed.password or '',
            'database': parsed.path.lstrip('/') if parsed.path else 'insurance_db',
            'pool_name': 'insurance_pool',
            'pool_size': 5,
            'pool_reset_session': True,
            'autocommit': False
        }
    else:
        # Use individual environment variables
        return {
            'host': os.getenv('MYSQL_HOST', 'localhost'),
            'port': int(os.getenv('MYSQL_PORT', 3306)),
            'user': os.getenv('MYSQL_USER', 'root'),
            'password': os.getenv('MYSQL_PASSWORD', ''),
            'database': os.getenv('MYSQL_DATABASE', 'insurance_db'),
            'pool_name': 'insurance_pool',
            'pool_size': 5,
            'pool_reset_session': True,
            'autocommit': False
        }

# MySQL connection pool configuration
DB_CONFIG = get_db_config()

# Create connection pool
try:
    connection_pool = pooling.MySQLConnectionPool(**DB_CONFIG)
    print(f"MySQL connection pool created successfully for database: {DB_CONFIG['database']}")
except Exception as e:
    print(f"Error creating connection pool: {e}")
    connection_pool = None


def get_connection():
    """Get database connection from pool"""
    if connection_pool:
        return connection_pool.get_connection()
    else:
        # Fallback to direct connection if pool fails
        return mysql.connector.connect(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            database=DB_CONFIG['database']
        )


def init_database():
    """Initialize database with all required tables"""
    conn = get_connection()
    cursor = conn.cursor()

    # Table 0: Users - stores user accounts
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INT PRIMARY KEY AUTO_INCREMENT,
            name VARCHAR(255) NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            is_admin BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP NULL,
            INDEX idx_users_email (email)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 1: Form Submissions - stores user form submissions with all details
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS form_submissions (
            id INT PRIMARY KEY AUTO_INCREMENT,
            user_id INT NOT NULL,
            user_name VARCHAR(255),
            user_email VARCHAR(255),

            -- Vehicle Information
            marque VARCHAR(100),
            modele VARCHAR(100),
            carburant VARCHAR(50),
            nombre_places INT,
            puissance_fiscale INT,
            date_mec VARCHAR(50),
            type_plaque VARCHAR(50),
            immatriculation VARCHAR(100),
            valeur_neuf DECIMAL(15,2) NOT NULL,
            valeur_actuelle DECIMAL(15,2) NOT NULL,

            -- Personal Information
            nom VARCHAR(100),
            prenom VARCHAR(100),
            telephone VARCHAR(50),
            email VARCHAR(255),
            date_naissance VARCHAR(50),
            date_permis VARCHAR(50),
            ville VARCHAR(100),
            agent_key VARCHAR(255),
            assureur_actuel VARCHAR(100),
            consent BOOLEAN,

            -- Metadata
            submission_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            ip_address VARCHAR(100),
            user_agent TEXT,

            FOREIGN KEY (user_id) REFERENCES users(id),
            INDEX idx_form_submissions_user (user_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 2: Scraper Results - stores results from scrapers for each form submission
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scraper_results (
            id INT PRIMARY KEY AUTO_INCREMENT,
            form_submission_id INT NOT NULL,
            user_id INT NOT NULL,
            provider_code VARCHAR(50) NOT NULL,
            provider_name VARCHAR(100) NOT NULL,

            -- Result data
            raw_response LONGTEXT,
            plan_count INT DEFAULT 0,
            fetch_time DECIMAL(10,3),
            status VARCHAR(50) DEFAULT 'success',
            error_message TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY (form_submission_id) REFERENCES form_submissions(id),
            FOREIGN KEY (user_id) REFERENCES users(id),
            INDEX idx_scraper_results_form (form_submission_id),
            INDEX idx_scraper_results_user (user_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 3: User Requests - stores user input parameters (kept for backward compatibility)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_requests (
            id INT PRIMARY KEY AUTO_INCREMENT,
            valeur_neuf DECIMAL(15,2) NOT NULL,
            valeur_venale DECIMAL(15,2) NOT NULL,
            request_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            ip_address VARCHAR(100),
            user_agent TEXT,
            total_fetch_time DECIMAL(10,3),
            status VARCHAR(50) DEFAULT 'pending',
            INDEX idx_requests_timestamp (request_timestamp)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')
    
    # Table 4: Provider Responses - stores raw response from each provider
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS provider_responses (
            id INT PRIMARY KEY AUTO_INCREMENT,
            request_id INT NOT NULL,
            provider_name VARCHAR(100) NOT NULL,
            provider_code VARCHAR(50) NOT NULL,
            raw_response LONGTEXT,
            fetch_time DECIMAL(10,3),
            status VARCHAR(50) DEFAULT 'success',
            error_message TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (request_id) REFERENCES user_requests(id),
            INDEX idx_responses_request (request_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 5: Insurance Plans - stores individual plan details
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS insurance_plans (
            id INT PRIMARY KEY AUTO_INCREMENT,
            response_id INT NOT NULL,
            request_id INT NOT NULL,
            provider_name VARCHAR(100) NOT NULL,
            plan_name VARCHAR(255) NOT NULL,
            plan_code VARCHAR(100),

            -- Annual pricing
            prime_net_annual DECIMAL(15,2),
            taxes_annual DECIMAL(15,2),
            prime_total_annual DECIMAL(15,2),

            -- Semi-annual pricing (6 months)
            prime_net_semi_annual DECIMAL(15,2),
            taxes_semi_annual DECIMAL(15,2),
            prime_total_semi_annual DECIMAL(15,2),

            -- Additional fees
            cnpac DECIMAL(15,2) DEFAULT 0,
            accessoires DECIMAL(15,2) DEFAULT 0,
            timbre DECIMAL(15,2) DEFAULT 0,

            -- Metadata
            is_promoted BOOLEAN DEFAULT 0,
            is_eligible BOOLEAN DEFAULT 1,
            color VARCHAR(50),
            plan_order INT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (response_id) REFERENCES provider_responses(id),
            FOREIGN KEY (request_id) REFERENCES user_requests(id),
            INDEX idx_plans_response (response_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')
    
    # Table 6: Plan Guarantees - stores guarantees/coverages for each plan
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS plan_guarantees (
            id INT PRIMARY KEY AUTO_INCREMENT,
            plan_id INT NOT NULL,
            guarantee_name VARCHAR(255) NOT NULL,
            guarantee_code VARCHAR(100),
            description TEXT,

            -- Coverage details
            capital_guarantee DECIMAL(15,2),
            franchise VARCHAR(100),
            prime_annual DECIMAL(15,2) DEFAULT 0,

            -- Status flags
            is_included BOOLEAN DEFAULT 1,
            is_obligatory BOOLEAN DEFAULT 0,
            is_optional BOOLEAN DEFAULT 0,

            -- For selectable options (like Bris de glace amounts)
            has_options BOOLEAN DEFAULT 0,
            options_json TEXT,
            selected_option VARCHAR(100),

            display_order INT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (plan_id) REFERENCES insurance_plans(id),
            INDEX idx_guarantees_plan (plan_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 7: Selectable Fields - stores selectable field definitions for plans
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS selectable_fields (
            id INT PRIMARY KEY AUTO_INCREMENT,
            plan_id INT NOT NULL,
            field_name VARCHAR(100) NOT NULL,
            field_title VARCHAR(255),
            field_order INT DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (plan_id) REFERENCES insurance_plans(id),
            INDEX idx_fields_plan (plan_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 8: Selectable Options - stores options for each selectable field
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS selectable_options (
            id INT PRIMARY KEY AUTO_INCREMENT,
            field_id INT NOT NULL,
            option_id VARCHAR(100) NOT NULL,
            option_label VARCHAR(255) NOT NULL,
            is_default BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (field_id) REFERENCES selectable_fields(id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 9: Option Combinations Pricing - stores pricing for different option combinations
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS option_combinations (
            id INT PRIMARY KEY AUTO_INCREMENT,
            plan_id INT NOT NULL,
            combination_key VARCHAR(255) NOT NULL,
            combination_params TEXT,

            -- Annual pricing
            prime_net_annual DECIMAL(15,2),
            taxes_annual DECIMAL(15,2),
            prime_total_annual DECIMAL(15,2),

            -- Semi-annual pricing
            prime_net_semi_annual DECIMAL(15,2),
            taxes_semi_annual DECIMAL(15,2),
            prime_total_semi_annual DECIMAL(15,2),

            is_default BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (plan_id) REFERENCES insurance_plans(id),
            INDEX idx_combinations_plan (plan_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 10: User Settings - stores user-specific settings for PDF generation
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_settings (
            id INT PRIMARY KEY AUTO_INCREMENT,
            user_id INT NOT NULL UNIQUE,
            company_name VARCHAR(255),
            logo_filename VARCHAR(255),
            footer_text TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            INDEX idx_user_settings_user (user_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Table 11: Scraper Settings - stores admin settings for enabling/disabling scrapers
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scraper_settings (
            id INT PRIMARY KEY AUTO_INCREMENT,
            scraper_code VARCHAR(50) NOT NULL UNIQUE,
            scraper_name VARCHAR(100) NOT NULL,
            is_enabled BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_scraper_settings_code (scraper_code)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    # Initialize default scrapers if table is empty
    cursor.execute('SELECT COUNT(*) FROM scraper_settings')
    result = cursor.fetchone()
    if result[0] == 0:
        default_scrapers = [
            ('axa', 'AXA Assurance', 1),
            ('sanlam', 'Sanlam', 1),
            ('mcma', 'MCMA (MAMDA)', 1),
            ('rma', 'RMA Watanya', 1)
        ]
        cursor.executemany('''
            INSERT INTO scraper_settings (scraper_code, scraper_name, is_enabled)
            VALUES (%s, %s, %s)
        ''', default_scrapers)

    # Table 12: API Keys - stores API keys for external access
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS api_keys (
            id INT PRIMARY KEY AUTO_INCREMENT,
            api_key VARCHAR(255) NOT NULL UNIQUE,
            description TEXT,
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_used TIMESTAMP NULL,
            created_by INT,
            FOREIGN KEY (created_by) REFERENCES users(id),
            INDEX idx_api_keys_key (api_key)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    ''')

    conn.commit()
    cursor.close()
    conn.close()
    print("Database initialized successfully")


class DatabaseManager:
    """Manager class for database operations"""

    # ============ User Management ============
    @staticmethod
    def create_user(name: str, email: str, password: str, is_admin: bool = False) -> int:
        """Create a new user and return its ID"""
        import hashlib
        conn = get_connection()
        cursor = conn.cursor()

        # Hash the password
        password_hash = hashlib.sha256(password.encode()).hexdigest()

        try:
            cursor.execute('''
                INSERT INTO users (name, email, password, is_admin)
                VALUES (%s, %s, %s, %s)
            ''', (name, email, password_hash, is_admin))

            user_id = cursor.lastrowid
            conn.commit()
            cursor.close()
            conn.close()
            return user_id
        except mysql.connector.IntegrityError:
            cursor.close()
            conn.close()
            return None  # Email already exists

    @staticmethod
    def get_user_by_email(email: str) -> Optional[Dict]:
        """Get user by email"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('SELECT * FROM users WHERE email = %s', (email,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        return row if row else None

    @staticmethod
    def verify_user(email: str, password: str) -> Optional[Dict]:
        """Verify user credentials and return user data if valid"""
        import hashlib
        password_hash = hashlib.sha256(password.encode()).hexdigest()

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('SELECT * FROM users WHERE email = %s AND password = %s', (email, password_hash))
        row = cursor.fetchone()

        if row:
            # Update last login
            update_cursor = conn.cursor()
            update_cursor.execute('UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = %s', (row['id'],))
            conn.commit()
            update_cursor.close()

        cursor.close()
        conn.close()
        return row if row else None

    @staticmethod
    def get_all_users(exclude_admin: bool = False) -> List[Dict]:
        """Get all users"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        if exclude_admin:
            cursor.execute('SELECT id, name, email, created_at, last_login FROM users WHERE is_admin = 0 ORDER BY created_at DESC')
        else:
            cursor.execute('SELECT id, name, email, is_admin, created_at, last_login FROM users ORDER BY created_at DESC')

        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def delete_user(user_id: int) -> bool:
        """Delete a user by ID (cannot delete admin users)"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        try:
            # Check if user is admin
            cursor.execute('SELECT is_admin FROM users WHERE id = %s', (user_id,))
            user = cursor.fetchone()

            if not user:
                cursor.close()
                conn.close()
                return False  # User not found

            if user['is_admin']:
                cursor.close()
                conn.close()
                return False  # Cannot delete admin users

            # Delete user
            cursor.execute('DELETE FROM users WHERE id = %s', (user_id,))
            conn.commit()
            deleted = cursor.rowcount > 0
            cursor.close()
            conn.close()
            return deleted
        except Exception as e:
            cursor.close()
            conn.close()
            return False

    # ============ User Settings Management ============
    @staticmethod
    def get_user_settings(user_id: int) -> Optional[Dict]:
        """Get user settings by user ID"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('SELECT * FROM user_settings WHERE user_id = %s', (user_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        return row if row else None

    @staticmethod
    def save_user_settings(user_id: int, company_name: str = None, logo_filename: str = None, footer_text: str = None) -> bool:
        """Save or update user settings"""
        conn = get_connection()
        cursor = conn.cursor()

        try:
            # Check if settings exist
            cursor.execute('SELECT id FROM user_settings WHERE user_id = %s', (user_id,))
            existing = cursor.fetchone()

            if existing:
                # Update existing
                cursor.execute('''
                    UPDATE user_settings
                    SET company_name = %s, logo_filename = %s, footer_text = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = %s
                ''', (company_name, logo_filename, footer_text, user_id))
            else:
                # Insert new
                cursor.execute('''
                    INSERT INTO user_settings (user_id, company_name, logo_filename, footer_text)
                    VALUES (%s, %s, %s, %s)
                ''', (user_id, company_name, logo_filename, footer_text))

            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            cursor.close()
            conn.close()
            return False

    @staticmethod
    def update_user_logo(user_id: int, logo_filename: str) -> bool:
        """Update only the logo filename for a user"""
        conn = get_connection()
        cursor = conn.cursor()

        try:
            # Check if settings exist
            cursor.execute('SELECT id FROM user_settings WHERE user_id = %s', (user_id,))
            existing = cursor.fetchone()

            if existing:
                cursor.execute('''
                    UPDATE user_settings SET logo_filename = %s, updated_at = CURRENT_TIMESTAMP WHERE user_id = %s
                ''', (logo_filename, user_id))
            else:
                cursor.execute('''
                    INSERT INTO user_settings (user_id, logo_filename) VALUES (%s, %s)
                ''', (user_id, logo_filename))

            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            cursor.close()
            conn.close()
            return False

    # ============ Scraper Settings Management ============
    @staticmethod
    def get_all_scrapers() -> List[Dict]:
        """Get all scrapers with their enabled status"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM scraper_settings ORDER BY scraper_code')
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def get_enabled_scrapers() -> List[str]:
        """Get list of enabled scraper codes"""
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT scraper_code FROM scraper_settings WHERE is_enabled = 1')
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [row[0] for row in rows]

    @staticmethod
    def is_scraper_enabled(scraper_code: str) -> bool:
        """Check if a specific scraper is enabled"""
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT is_enabled FROM scraper_settings WHERE scraper_code = %s', (scraper_code,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return bool(row[0]) if row else False

    @staticmethod
    def toggle_scraper(scraper_code: str, is_enabled: bool) -> bool:
        """Enable or disable a scraper"""
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE scraper_settings
                SET is_enabled = %s, updated_at = CURRENT_TIMESTAMP
                WHERE scraper_code = %s
            ''', (is_enabled, scraper_code))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            print(f"Error toggling scraper {scraper_code}: {e}")
            cursor.close()
            conn.close()
            return False

    # ============ API Key Management ============
    @staticmethod
    def create_api_key(description: str = None, created_by: int = None) -> Optional[str]:
        """Create a new API key and return it"""
        import secrets
        conn = get_connection()
        cursor = conn.cursor()

        # Generate a secure random API key
        api_key = secrets.token_urlsafe(32)

        try:
            cursor.execute('''
                INSERT INTO api_keys (api_key, description, created_by)
                VALUES (%s, %s, %s)
            ''', (api_key, description, created_by))
            conn.commit()
            cursor.close()
            conn.close()
            return api_key
        except Exception as e:
            print(f"Error creating API key: {e}")
            cursor.close()
            conn.close()
            return None

    @staticmethod
    def validate_api_key(api_key: str) -> bool:
        """Validate an API key and update last_used timestamp"""
        conn = get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute('''
                SELECT is_active FROM api_keys WHERE api_key = %s
            ''', (api_key,))
            result = cursor.fetchone()

            if result and result[0]:  # Key exists and is active
                # Update last_used timestamp
                cursor.execute('''
                    UPDATE api_keys SET last_used = CURRENT_TIMESTAMP WHERE api_key = %s
                ''', (api_key,))
                conn.commit()
                cursor.close()
                conn.close()
                return True
            else:
                cursor.close()
                conn.close()
                return False
        except Exception as e:
            print(f"Error validating API key: {e}")
            cursor.close()
            conn.close()
            return False

    @staticmethod
    def get_all_api_keys() -> List[Dict]:
        """Get all API keys (admin only)"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, api_key, description, is_active, created_at, last_used FROM api_keys ORDER BY created_at DESC')
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def toggle_api_key(api_key: str, is_active: bool) -> bool:
        """Enable or disable an API key"""
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE api_keys SET is_active = %s WHERE api_key = %s
            ''', (is_active, api_key))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            print(f"Error toggling API key: {e}")
            cursor.close()
            conn.close()
            return False

    @staticmethod
    def delete_api_key(api_key: str) -> bool:
        """Delete an API key"""
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM api_keys WHERE api_key = %s', (api_key,))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            print(f"Error deleting API key: {e}")
            cursor.close()
            conn.close()
            return False

    # ============ Form Submission Management ============
    @staticmethod
    def save_form_submission(user_id: int, form_data: dict, ip_address: str = None, user_agent: str = None, user_name: str = None, user_email: str = None) -> int:
        """Save form submission and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO form_submissions
            (user_id, user_name, user_email, marque, modele, carburant, nombre_places, puissance_fiscale,
             date_mec, type_plaque, immatriculation, valeur_neuf, valeur_actuelle,
             nom, prenom, telephone, email, date_naissance, date_permis, ville,
             agent_key, assureur_actuel, consent, ip_address, user_agent)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            user_id,
            user_name,
            user_email,
            form_data.get('marque'),
            form_data.get('modele'),
            form_data.get('carburant'),
            form_data.get('nombre_places'),
            form_data.get('puissance_fiscale'),
            form_data.get('date_mec'),
            form_data.get('type_plaque'),
            form_data.get('immatriculation'),
            form_data.get('valeur_neuf'),
            form_data.get('valeur_actuelle'),
            form_data.get('nom'),
            form_data.get('prenom'),
            form_data.get('telephone'),
            form_data.get('email'),
            form_data.get('date_naissance'),
            form_data.get('date_permis'),
            form_data.get('ville'),
            form_data.get('agent_key'),
            form_data.get('assureur_actuel'),
            form_data.get('consent'),
            ip_address,
            user_agent
        ))

        submission_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return submission_id

    @staticmethod
    def save_scraper_result(form_submission_id: int, user_id: int, provider_code: str,
                           provider_name: str, raw_response: dict, plan_count: int,
                           fetch_time: float, status: str = 'success', error_message: str = None) -> int:
        """Save scraper result and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO scraper_results
            (form_submission_id, user_id, provider_code, provider_name, raw_response,
             plan_count, fetch_time, status, error_message)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            form_submission_id, user_id, provider_code, provider_name,
            json.dumps(raw_response, ensure_ascii=False) if raw_response else None,
            plan_count, fetch_time, status, error_message
        ))

        result_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return result_id

    @staticmethod
    def get_user_submissions(user_id: int, limit: int = 50) -> List[Dict]:
        """Get form submissions for a user"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('''
            SELECT * FROM form_submissions
            WHERE user_id = %s
            ORDER BY submission_timestamp DESC
            LIMIT %s
        ''', (user_id, limit))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def create_request(valeur_neuf: float, valeur_venale: float,
                       ip_address: str = None, user_agent: str = None) -> int:
        """Create a new user request and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO user_requests (valeur_neuf, valeur_venale, ip_address, user_agent)
            VALUES (%s, %s, %s, %s)
        ''', (valeur_neuf, valeur_venale, ip_address, user_agent))
        
        request_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return request_id
    
    @staticmethod
    def update_request_status(request_id: int, status: str, fetch_time: float = None):
        """Update request status and fetch time"""
        conn = get_connection()
        cursor = conn.cursor()
        
        if fetch_time:
            cursor.execute('''
                UPDATE user_requests SET status = %s, total_fetch_time = %s WHERE id = %s
            ''', (status, fetch_time, request_id))
        else:
            cursor.execute('''
                UPDATE user_requests SET status = %s WHERE id = %s
            ''', (status, request_id))
        
        conn.commit()
        cursor.close()
        conn.close()

    @staticmethod
    def save_provider_response(request_id: int, provider_name: str, provider_code: str,
                                raw_response: dict, fetch_time: float, 
                                status: str = 'success', error_message: str = None) -> int:
        """Save provider response and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO provider_responses 
            (request_id, provider_name, provider_code, raw_response, fetch_time, status, error_message)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (request_id, provider_name, provider_code, 
              json.dumps(raw_response, ensure_ascii=False) if raw_response else None,
              fetch_time, status, error_message))
        
        response_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return response_id
    
    @staticmethod
    def save_insurance_plan(response_id: int, request_id: int, plan_data: dict) -> int:
        """Save insurance plan and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO insurance_plans 
            (response_id, request_id, provider_name, plan_name, plan_code,
             prime_net_annual, taxes_annual, prime_total_annual,
             prime_net_semi_annual, taxes_semi_annual, prime_total_semi_annual,
             cnpac, accessoires, timbre, is_promoted, is_eligible, color, plan_order)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            response_id, request_id,
            plan_data.get('provider_name'),
            plan_data.get('plan_name'),
            plan_data.get('plan_code'),
            plan_data.get('prime_net_annual'),
            plan_data.get('taxes_annual'),
            plan_data.get('prime_total_annual'),
            plan_data.get('prime_net_semi_annual'),
            plan_data.get('taxes_semi_annual'),
            plan_data.get('prime_total_semi_annual'),
            plan_data.get('cnpac', 0),
            plan_data.get('accessoires', 0),
            plan_data.get('timbre', 0),
            plan_data.get('is_promoted', False),
            plan_data.get('is_eligible', True),
            plan_data.get('color'),
            plan_data.get('plan_order', 0)
        ))
        
        plan_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return plan_id
    
    @staticmethod
    def save_plan_guarantee(plan_id: int, guarantee_data: dict) -> int:
        """Save plan guarantee and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO plan_guarantees 
            (plan_id, guarantee_name, guarantee_code, description,
             capital_guarantee, franchise, prime_annual,
             is_included, is_obligatory, is_optional,
             has_options, options_json, selected_option, display_order)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            plan_id,
            guarantee_data.get('guarantee_name'),
            guarantee_data.get('guarantee_code'),
            guarantee_data.get('description'),
            guarantee_data.get('capital_guarantee'),
            guarantee_data.get('franchise'),
            guarantee_data.get('prime_annual', 0),
            guarantee_data.get('is_included', True),
            guarantee_data.get('is_obligatory', False),
            guarantee_data.get('is_optional', False),
            guarantee_data.get('has_options', False),
            json.dumps(guarantee_data.get('options', [])) if guarantee_data.get('options') else None,
            guarantee_data.get('selected_option'),
            guarantee_data.get('display_order', 0)
        ))
        
        guarantee_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return guarantee_id
    
    @staticmethod
    def save_selectable_field(plan_id: int, field_name: str, field_title: str, field_order: int = 0) -> int:
        """Save selectable field and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO selectable_fields (plan_id, field_name, field_title, field_order)
            VALUES (%s, %s, %s, %s)
        ''', (plan_id, field_name, field_title, field_order))

        field_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return field_id

    @staticmethod
    def save_selectable_option(field_id: int, option_id: str, option_label: str, is_default: bool = False) -> int:
        """Save selectable option and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO selectable_options (field_id, option_id, option_label, is_default)
            VALUES (%s, %s, %s, %s)
        ''', (field_id, option_id, option_label, is_default))

        option_id_pk = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return option_id_pk

    @staticmethod
    def save_option_combination(plan_id: int, combination_key: str, combination_params: str,
                               pricing_data: dict, is_default: bool = False) -> int:
        """Save option combination pricing and return its ID"""
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO option_combinations
            (plan_id, combination_key, combination_params,
             prime_net_annual, taxes_annual, prime_total_annual,
             prime_net_semi_annual, taxes_semi_annual, prime_total_semi_annual,
             is_default)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            plan_id, combination_key, combination_params,
            pricing_data.get('prime_net_annual'),
            pricing_data.get('taxes_annual'),
            pricing_data.get('prime_total_annual'),
            pricing_data.get('prime_net_semi_annual'),
            pricing_data.get('taxes_semi_annual'),
            pricing_data.get('prime_total_semi_annual'),
            is_default
        ))

        combination_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return combination_id

    @staticmethod
    def get_option_combinations(plan_id: int) -> List[Dict]:
        """Get all option combinations for a plan"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('''
            SELECT * FROM option_combinations
            WHERE plan_id = %s
            ORDER BY is_default DESC
        ''', (plan_id,))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def get_request_history(limit: int = 50) -> List[Dict]:
        """Get recent request history"""
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute('''
            SELECT * FROM user_requests
            ORDER BY request_timestamp DESC
            LIMIT %s
        ''', (limit,))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    @staticmethod
    def export_database_to_excel(filepath: str) -> bool:
        """Export entire database to Excel file with each table as a sheet"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            # Get list of all tables
            cursor.execute("""
                SELECT TABLE_NAME
                FROM information_schema.tables
                WHERE table_schema = %s
                ORDER BY TABLE_NAME
            """, (DB_CONFIG['database'],))

            tables = [row['TABLE_NAME'] for row in cursor.fetchall()]

            # Create Excel workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            for table_name in tables:
                # Get table data
                cursor.execute(f"SELECT * FROM `{table_name}`")
                rows = cursor.fetchall()

                if not rows:
                    continue  # Skip empty tables

                # Create sheet for this table
                ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit is 31 chars

                # Get column names
                columns = list(rows[0].keys())

                # Write header row
                ws.append(columns)

                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Write data rows
                for row in rows:
                    ws.append(list(row.values()))

                # Auto-adjust column widths
                for idx, column in enumerate(columns, 1):
                    column_letter = get_column_letter(idx)
                    max_length = len(str(column))

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                        for cell in row:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(filepath)

            cursor.close()
            conn.close()
            return True

        except Exception as e:
            print(f"Error exporting database to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False


# Initialize database on module import
if __name__ == "__main__":
    init_database()
